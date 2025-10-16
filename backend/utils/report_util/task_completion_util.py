from typing import List, Dict, Any
from datetime import date, datetime
from io import BytesIO
from backend.wrappers.supabase_wrapper.supabase_crud import SupabaseCRUD
from backend.schemas.report_schemas import TaskCompletionResponse, TaskCompletionItem
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


class TaskCompletionReportGenerator:
    """
    Generates task completion reports filtered by project or staff member.
    """

    def __init__(self):
        self.crud = SupabaseCRUD()

    def generate_report(
        self,
        scope_type: str,
        scope_id: str,
        start_date: date,
        end_date: date
    ) -> TaskCompletionResponse:
        """
        Generate task completion report based on scope and date range.

        Args:
            scope_type: "project" or "staff"
            scope_id: Project ID or Staff user UUID
            start_date: Start date for filtering tasks
            end_date: End date for filtering tasks

        Returns:
            TaskCompletionResponse with filtered tasks
        """
        # Get scope name (project name or staff email)
        scope_name = self._get_scope_name(scope_type, scope_id)

        # Get filtered tasks
        if scope_type == "project":
            tasks = self._get_tasks_by_project(scope_id, start_date, end_date)
        else:  # staff
            tasks = self._get_tasks_by_staff(scope_id, start_date, end_date)

        # Convert to response format
        task_items = [self._convert_to_task_item(task) for task in tasks]

        return TaskCompletionResponse(
            scope_type=scope_type,
            scope_id=scope_id,
            scope_name=scope_name,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            total_tasks=len(task_items),
            tasks=task_items
        )

    def _get_scope_name(self, scope_type: str, scope_id: str) -> str:
        """Get the display name for the scope (project name or staff email)"""
        if scope_type == "project":
            projects = self.crud.select("projects", filters={"id": scope_id})
            if projects:
                return projects[0].get("name", "Unknown Project")
            return "Unknown Project"
        else:  # staff
            users = self.crud.select("users", filters={"uuid": scope_id})
            if users:
                return users[0].get("email", "Unknown User")
            return "Unknown User"

    def _get_tasks_by_project(
        self,
        project_id: str,
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Get all tasks for a specific project within date range"""
        all_tasks = self.crud.select("tasks", filters={"project_id": project_id})
        return self._filter_by_date_range(all_tasks, start_date, end_date)

    def _get_tasks_by_staff(
        self,
        user_id: str,
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Get all tasks assigned to a specific staff member within date range"""
        all_tasks = self.crud.select("tasks")

        # Filter tasks where user is in assignee_ids
        assigned_tasks = [
            task for task in all_tasks
            if user_id in task.get("assignee_ids", [])
        ]

        return self._filter_by_date_range(assigned_tasks, start_date, end_date)

    def _filter_by_date_range(
        self,
        tasks: List[Dict[str, Any]],
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Filter tasks by due date range"""
        filtered_tasks = []
        for task in tasks:
            if task.get("due_date"):
                task_due_date = datetime.fromisoformat(task["due_date"]).date()
                if start_date <= task_due_date <= end_date:
                    filtered_tasks.append(task)
        return filtered_tasks

    def _convert_to_task_item(self, task: Dict[str, Any]) -> TaskCompletionItem:
        """Convert task data to TaskCompletionItem"""
        due_date_str = task.get("due_date", "")
        status = task.get("status", "TO_DO")

        # Calculate overdue: task is overdue if due_date is past and status is not COMPLETED
        is_overdue = False
        if due_date_str:
            task_due_date = datetime.fromisoformat(due_date_str).date()
            is_overdue = task_due_date < date.today() and status != "COMPLETED"

        return TaskCompletionItem(
            task_title=task.get("title", "Untitled"),
            priority=task.get("priority", 5),
            status=status,
            due_date=due_date_str,
            overdue=is_overdue
        )

    def generate_excel_bytes(
        self,
        scope_type: str,
        scope_id: str,
        scope_name: str,
        start_date: date,
        end_date: date,
        tasks: List[TaskCompletionItem]
    ) -> BytesIO:
        """
        Generate Excel file in memory.

        Returns:
            BytesIO object containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Task Completion Report"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Report metadata
        ws["A1"] = "Task Completion Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Scope: {scope_type.upper()} - {scope_name}"
        ws["A3"] = f"Period: {start_date.isoformat()} to {end_date.isoformat()}"
        ws["A4"] = f"Total Tasks: {len(tasks)}"

        # Column headers (row 6)
        headers = ["Task Title", "Priority", "Status", "Due Date", "Overdue"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Data rows
        for row_num, task in enumerate(tasks, start=7):
            ws.cell(row=row_num, column=1, value=task.task_title)
            ws.cell(row=row_num, column=2, value=task.priority).alignment = center_alignment
            ws.cell(row=row_num, column=3, value=task.status)
            ws.cell(row=row_num, column=4, value=task.due_date)
            ws.cell(row=row_num, column=5, value="Yes" if task.overdue else "No").alignment = center_alignment

        # Adjust column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 10
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 10

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_pdf_bytes(
        self,
        scope_type: str,
        scope_id: str,
        scope_name: str,
        start_date: date,
        end_date: date,
        tasks: List[TaskCompletionItem]
    ) -> BytesIO:
        """
        Generate PDF file in memory.

        Returns:
            BytesIO object containing the PDF file
        """
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("Task Completion Report", styles["Title"])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Metadata
        metadata_style = styles["Normal"]
        elements.append(Paragraph(f"<b>Scope:</b> {scope_type.upper()} - {scope_name}", metadata_style))
        elements.append(Paragraph(f"<b>Period:</b> {start_date.isoformat()} to {end_date.isoformat()}", metadata_style))
        elements.append(Paragraph(f"<b>Total Tasks:</b> {len(tasks)}", metadata_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Table data
        table_data = [["Task Title", "Priority", "Status", "Due Date", "Overdue"]]
        for task in tasks:
            table_data.append([
                task.task_title[:40],  # Truncate long titles
                str(task.priority),
                task.status,
                task.due_date,
                "Yes" if task.overdue else "No"
            ])

        # Create table
        table = Table(table_data, colWidths=[3 * inch, 0.7 * inch, 1 * inch, 1 * inch, 0.7 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),  # Left-align task titles
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output
