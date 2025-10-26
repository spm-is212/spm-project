from typing import List, Dict, Any
from datetime import date, datetime
from io import BytesIO
from backend.wrappers.supabase_wrapper.supabase_crud import SupabaseCRUD
from backend.schemas.report_schemas import LoggedTimeResponse, LoggedTimeItem
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


class LoggedTimeReportGenerator:
    """
    Generates logged time reports showing time logged per user and task.
    """

    def __init__(self):
        self.crud = SupabaseCRUD()

    def generate_report(
        self,
        scope_type: str,
        scope_id: str,
        start_date: date,
        end_date: date
    ) -> LoggedTimeResponse:
        """
        Generate logged time report based on scope and date range.

        Args:
            scope_type: "department" or "project"
            scope_id: Department name or Project ID
            start_date: Start date for filtering tasks
            end_date: End date for filtering tasks

        Returns:
            LoggedTimeResponse with time entries
        """
        # Get scope name
        scope_name = self._get_scope_name(scope_type, scope_id)

        # Get time entries
        if scope_type == "department":
            time_entries = self._get_entries_by_department(scope_id, start_date, end_date)
        else:  # project
            time_entries = self._get_entries_by_project(scope_id, start_date, end_date)

        # Calculate total hours - FIXED: Handle None values
        total_hours = sum(entry.time_log if entry.time_log is not None else 0.0 for entry in time_entries)

        return LoggedTimeResponse(
            scope_type=scope_type,
            scope_id=scope_id,
            scope_name=scope_name,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            total_entries=len(time_entries),
            total_hours=round(total_hours, 2),
            time_entries=time_entries
        )

    def _get_scope_name(self, scope_type: str, scope_id: str) -> str:
        """Get the display name for the scope"""
        if scope_type == "department":
            return scope_id  # Department name is the ID itself
        else:  # project
            projects = self.crud.select("projects", filters={"id": scope_id})
            if projects:
                return projects[0].get("name", "Unknown Project")
            return "Unknown Project"

    def _get_entries_by_department(
        self,
        department_name: str,
        start_date: date,
        end_date: date
    ) -> List[LoggedTimeItem]:
        """Get time entries for all staff in a department"""
        # Get all users in this department
        all_users = self.crud.select("users")
        department_users = [
            user for user in all_users
            if department_name.lower() in [dept.lower() for dept in user.get("departments", [])]
        ]

        # Get all tasks
        all_tasks = self.crud.select("tasks")
        filtered_tasks = self._filter_by_date_range(all_tasks, start_date, end_date)

        # Build time entries for each user-task combination
        time_entries = []
        for user in department_users:
            user_id = user.get("uuid")
            user_email = user.get("email", "Unknown")

            # Get tasks where this user is assigned
            user_tasks = [
                task for task in filtered_tasks
                if user_id in task.get("assignee_ids", [])
            ]

            for task in user_tasks:
                entry = self._create_time_entry(user_email, task)
                time_entries.append(entry)

        return time_entries

    def _get_entries_by_project(
        self,
        project_id: str,
        start_date: date,
        end_date: date
    ) -> List[LoggedTimeItem]:
        """Get time entries for all staff in a project"""
        # Get all tasks for this project
        project_tasks = self.crud.select("tasks", filters={"project_id": project_id})
        filtered_tasks = self._filter_by_date_range(project_tasks, start_date, end_date)

        # Get user details
        all_users = self.crud.select("users")
        user_map = {user["uuid"]: user.get("email", "Unknown") for user in all_users}

        # Build time entries
        time_entries = []
        for task in filtered_tasks:
            for user_id in task.get("assignee_ids", []):
                user_email = user_map.get(user_id, "Unknown User")
                entry = self._create_time_entry(user_email, task)
                time_entries.append(entry)

        return time_entries

    def _filter_by_date_range(
        self,
        tasks: List[Dict[str, Any]],
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Filter tasks by due date range"""
        filtered_tasks = []
        for task in tasks:
            if task.get("due_date"):
                task_due_date = datetime.fromisoformat(task["due_date"]).date()
                if start_date <= task_due_date <= end_date:
                    filtered_tasks.append(task)
        return filtered_tasks

    def _create_time_entry(
        self,
        staff_email: str,
        task: Dict[str, Any]
    ) -> LoggedTimeItem:
        """Create a LoggedTimeItem from task data"""
        due_date_str = task.get("due_date", "")
        status = task.get("status", "TO_DO")
        
        # FIXED: Handle None/null time_log values
        time_log_value = task.get("time_log")
        if time_log_value is None:
            time_log = 0.0
        else:
            try:
                time_log = float(time_log_value)
            except (ValueError, TypeError):
                time_log = 0.0

        # Calculate overdue
        is_overdue = False
        if due_date_str:
            task_due_date = datetime.fromisoformat(due_date_str).date()
            is_overdue = task_due_date < date.today() and status != "COMPLETED"

        return LoggedTimeItem(
            staff_name=staff_email,
            task_title=task.get("title", "Untitled"),
            time_log=round(time_log, 2),
            status=status,
            due_date=due_date_str,
            overdue=is_overdue
        )

    def generate_excel_bytes(
        self,
        scope_type: str,
        scope_name: str,
        start_date: date,
        end_date: date,
        time_entries: List[LoggedTimeItem],
        total_hours: float
    ) -> BytesIO:
        """Generate Excel file in memory"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Logged Time Report"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Report metadata
        ws["A1"] = "Logged Time Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Scope: {scope_type.upper()} - {scope_name}"
        ws["A3"] = f"Period: {start_date.isoformat()} to {end_date.isoformat()}"
        ws["A4"] = f"Total Hours: {total_hours}"
        ws["A5"] = f"Total Entries: {len(time_entries)}"

        # Column headers (row 7)
        headers = ["Staff Name", "Task Title", "Time Log (hrs)", "Status", "Due Date", "Overdue"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Data rows
        for row_num, entry in enumerate(time_entries, start=8):
            ws.cell(row=row_num, column=1, value=entry.staff_name)
            ws.cell(row=row_num, column=2, value=entry.task_title)
            ws.cell(row=row_num, column=3, value=entry.time_log).alignment = center_alignment
            ws.cell(row=row_num, column=4, value=entry.status)
            ws.cell(row=row_num, column=5, value=entry.due_date)
            ws.cell(row=row_num, column=6, value="Yes" if entry.overdue else "No").alignment = center_alignment

        # Adjust column widths
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 35
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 10

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_pdf_bytes(
        self,
        scope_type: str,
        scope_name: str,
        start_date: date,
        end_date: date,
        time_entries: List[LoggedTimeItem],
        total_hours: float
    ) -> BytesIO:
        """Generate PDF file in memory"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("Logged Time Report", styles["Title"])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Metadata
        metadata_style = styles["Normal"]
        elements.append(Paragraph(f"<b>Scope:</b> {scope_type.upper()} - {scope_name}", metadata_style))
        elements.append(Paragraph(f"<b>Period:</b> {start_date.isoformat()} to {end_date.isoformat()}", metadata_style))
        elements.append(Paragraph(f"<b>Total Hours:</b> {total_hours}", metadata_style))
        elements.append(Paragraph(f"<b>Total Entries:</b> {len(time_entries)}", metadata_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Table data
        table_data = [["Staff Name", "Task Title", "Time (hrs)", "Status", "Due Date", "Overdue"]]
        for entry in time_entries:
            table_data.append([
                entry.staff_name[:25],
                entry.task_title[:30],
                str(entry.time_log),
                entry.status,
                entry.due_date,
                "Yes" if entry.overdue else "No"
            ])

        # Create table
        table = Table(table_data, colWidths=[2*inch, 2.5*inch, 0.8*inch, 1*inch, 1*inch, 0.7*inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (1, -1), "LEFT"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 9),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output