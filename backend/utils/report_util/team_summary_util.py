from typing import List, Dict, Any
from datetime import date, datetime
from io import BytesIO
from backend.wrappers.supabase_wrapper.supabase_crud import SupabaseCRUD
from backend.schemas.report_schemas import TeamSummaryResponse, StaffTaskSummary
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch


class TeamSummaryReportGenerator:
    """
    Generates team summary reports showing task counts by status per staff member.
    """

    def __init__(self):
        self.crud = SupabaseCRUD()

    def generate_report(
        self,
        scope_type: str,
        scope_id: str,
        time_frame: str,
        start_date: date,
        end_date: date
    ) -> TeamSummaryResponse:
        """
        Generate team summary report based on scope and date range.

        Args:
            scope_type: "department" or "project"
            scope_id: Department name or Project ID
            time_frame: "weekly" or "monthly"
            start_date: Start date for filtering tasks
            end_date: End date for filtering tasks

        Returns:
            TeamSummaryResponse with staff task summaries
        """
        # Get scope name
        scope_name = self._get_scope_name(scope_type, scope_id)

        # Get tasks and aggregate by staff
        if scope_type == "department":
            staff_summaries = self._get_summaries_by_department(scope_id, start_date, end_date)
        else:  # project
            staff_summaries = self._get_summaries_by_project(scope_id, start_date, end_date)

        return TeamSummaryResponse(
            scope_type=scope_type,
            scope_id=scope_id,
            scope_name=scope_name,
            time_frame=time_frame,
            start_date=start_date.isoformat(),
            end_date=end_date.isoformat(),
            total_staff=len(staff_summaries),
            staff_summaries=staff_summaries
        )

    def _get_scope_name(self, scope_type: str, scope_id: str) -> str:
        """Get the display name for the scope"""
        if scope_type == "department":
            return scope_id  # Department name is the ID itself
        else:  # project
            projects = self.crud.select("projects", filters={"id": scope_id})
            if projects:
                return projects[0].get("name", "Unknown Project")
            return "Unknown Project"

    def _get_summaries_by_department(
        self,
        department_name: str,
        start_date: date,
        end_date: date
    ) -> List[StaffTaskSummary]:
        """Get task summaries for all staff in a department"""
        # Get all users in this department
        all_users = self.crud.select("users")
        department_users = [
            user for user in all_users
            if department_name.lower() in [dept.lower() for dept in user.get("departments", [])]
        ]

        # Get all tasks
        all_tasks = self.crud.select("tasks")
        filtered_tasks = self._filter_by_date_range(all_tasks, start_date, end_date)

        # Aggregate tasks by user
        staff_summaries = []
        for user in department_users:
            user_id = user.get("uuid")
            user_email = user.get("email", "Unknown")

            # Get tasks where this user is assigned
            user_tasks = [
                task for task in filtered_tasks
                if user_id in task.get("assignee_ids", [])
            ]

            summary = self._aggregate_task_counts(user_email, user_tasks)
            if summary.total_tasks > 0:  # Only include staff with tasks
                staff_summaries.append(summary)

        return staff_summaries

    def _get_summaries_by_project(
        self,
        project_id: str,
        start_date: date,
        end_date: date
    ) -> List[StaffTaskSummary]:
        """Get task summaries for all staff in a project"""
        # Get all tasks for this project
        project_tasks = self.crud.select("tasks", filters={"project_id": project_id})
        filtered_tasks = self._filter_by_date_range(project_tasks, start_date, end_date)

        # Get unique assignee IDs from tasks
        assignee_ids = set()
        for task in filtered_tasks:
            assignee_ids.update(task.get("assignee_ids", []))

        # Get user details
        all_users = self.crud.select("users")
        user_map = {user["uuid"]: user.get("email", "Unknown") for user in all_users}

        # Aggregate tasks by user
        staff_summaries = []
        for user_id in assignee_ids:
            user_email = user_map.get(user_id, "Unknown User")

            # Get tasks where this user is assigned
            user_tasks = [
                task for task in filtered_tasks
                if user_id in task.get("assignee_ids", [])
            ]

            summary = self._aggregate_task_counts(user_email, user_tasks)
            staff_summaries.append(summary)

        return staff_summaries

    def _filter_by_date_range(
        self,
        tasks: List[Dict[str, Any]],
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Filter tasks by due date range"""
        filtered_tasks = []
        for task in tasks:
            if task.get("due_date"):
                task_due_date = datetime.fromisoformat(task["due_date"]).date()
                if start_date <= task_due_date <= end_date:
                    filtered_tasks.append(task)
        return filtered_tasks

    def _aggregate_task_counts(
        self,
        staff_name: str,
        tasks: List[Dict[str, Any]]
    ) -> StaffTaskSummary:
        """Aggregate task counts by status for a staff member"""
        blocked = 0
        in_progress = 0
        completed = 0
        overdue = 0

        today = date.today()

        for task in tasks:
            status = task.get("status", "TO_DO")
            due_date_str = task.get("due_date")

            # Check if overdue
            is_overdue = False
            if due_date_str:
                task_due_date = datetime.fromisoformat(due_date_str).date()
                is_overdue = task_due_date < today and status != "COMPLETED"

            # Count by status
            if is_overdue:
                overdue += 1
            elif status == "BLOCKED":
                blocked += 1
            elif status == "IN_PROGRESS":
                in_progress += 1
            elif status == "COMPLETED":
                completed += 1

        total = blocked + in_progress + completed + overdue

        return StaffTaskSummary(
            staff_name=staff_name,
            blocked=blocked,
            in_progress=in_progress,
            completed=completed,
            overdue=overdue,
            total_tasks=total
        )

    def generate_excel_bytes(
        self,
        scope_type: str,
        scope_name: str,
        time_frame: str,
        start_date: date,
        end_date: date,
        staff_summaries: List[StaffTaskSummary]
    ) -> BytesIO:
        """Generate Excel file in memory"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Team Summary Report"

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        center_alignment = Alignment(horizontal="center", vertical="center")

        # Report metadata
        ws["A1"] = "Team Summary Report"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Scope: {scope_type.upper()} - {scope_name}"
        ws["A3"] = f"Time Frame: {time_frame.upper()}"
        ws["A4"] = f"Period: {start_date.isoformat()} to {end_date.isoformat()}"
        ws["A5"] = f"Total Staff: {len(staff_summaries)}"

        # Column headers (row 7)
        headers = ["Staff Name", "Blocked", "In Progress", "Completed", "Overdue", "Total Tasks"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_alignment

        # Data rows
        for row_num, summary in enumerate(staff_summaries, start=8):
            ws.cell(row=row_num, column=1, value=summary.staff_name)
            ws.cell(row=row_num, column=2, value=summary.blocked).alignment = center_alignment
            ws.cell(row=row_num, column=3, value=summary.in_progress).alignment = center_alignment
            ws.cell(row=row_num, column=4, value=summary.completed).alignment = center_alignment
            ws.cell(row=row_num, column=5, value=summary.overdue).alignment = center_alignment
            ws.cell(row=row_num, column=6, value=summary.total_tasks).alignment = center_alignment

        # Adjust column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 12
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 15

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        return output

    def generate_pdf_bytes(
        self,
        scope_type: str,
        scope_name: str,
        time_frame: str,
        start_date: date,
        end_date: date,
        staff_summaries: List[StaffTaskSummary]
    ) -> BytesIO:
        """Generate PDF file in memory"""
        output = BytesIO()
        doc = SimpleDocTemplate(output, pagesize=A4)
        elements = []
        styles = getSampleStyleSheet()

        # Title
        title = Paragraph("Team Summary Report", styles["Title"])
        elements.append(title)
        elements.append(Spacer(1, 0.2 * inch))

        # Metadata
        metadata_style = styles["Normal"]
        elements.append(Paragraph(f"<b>Scope:</b> {scope_type.upper()} - {scope_name}", metadata_style))
        elements.append(Paragraph(f"<b>Time Frame:</b> {time_frame.upper()}", metadata_style))
        elements.append(Paragraph(f"<b>Period:</b> {start_date.isoformat()} to {end_date.isoformat()}", metadata_style))
        elements.append(Paragraph(f"<b>Total Staff:</b> {len(staff_summaries)}", metadata_style))
        elements.append(Spacer(1, 0.3 * inch))

        # Table data
        table_data = [["Staff Name", "Blocked", "In Progress", "Completed", "Overdue", "Total"]]
        for summary in staff_summaries:
            table_data.append([
                summary.staff_name[:30],  # Truncate long names
                str(summary.blocked),
                str(summary.in_progress),
                str(summary.completed),
                str(summary.overdue),
                str(summary.total_tasks)
            ])

        # Create table
        table = Table(table_data, colWidths=[2.5 * inch, 0.8 * inch, 1 * inch, 0.9 * inch, 0.8 * inch, 0.8 * inch])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#366092")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("ALIGN", (0, 1), (0, -1), "LEFT"),  # Left-align staff names
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
            ("BACKGROUND", (0, 1), (-1, -1), colors.beige),
            ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ("FONTSIZE", (0, 1), (-1, -1), 8),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.lightgrey])
        ]))

        elements.append(table)

        # Build PDF
        doc.build(elements)
        output.seek(0)
        return output
